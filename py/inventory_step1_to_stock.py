from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OUTPUT_SHEET_NAME = "stock"
SKU_SHEET_CANDIDATES = ["SKU", "SKU Mapping", "Legacy Mapping Product"]
TRANSIT_START_DATE = date(2026, 8, 1)
TRANSIT_END_DATE = date(2026, 12, 31)
ALLOC_SHEET_NAME = "To be allocated"

WH_CODE_MAP = {
    "SPNL": "NL",
    "SPFR": "FR",
    "SPTN": "FR",
    "SPUK": "UK",
    "SPIT": "IT",
    "SPES": "ES",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_lower(value: Any) -> str:
    return normalize_text(value).lower()


def safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def build_header_index(headers: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = normalize_lower(header)
        if key and key not in out:
            out[key] = idx
    return out


def map_country_code(sales_org_name: str) -> str:
    name = normalize_lower(sales_org_name)
    if "netherland" in name or "netherlands" in name:
        return "NL"
    if "france" in name:
        return "FR"
    if "united kingdom" in name or name == "uk" or name.endswith(" uk"):
        return "UK"
    if "italy" in name:
        return "IT"
    if "spain" in name:
        return "ES"
    return normalize_text(sales_org_name)[:2].upper()


def resolve_output_category(source_category: Any, family: Any) -> str:
    category_text = normalize_upper_category(source_category)
    if category_text == "HP":
        return "HP"

    family_text = normalize_lower(family)
    if family_text == "reserve":
        return "ESS"

    if category_text in {"ESS", "PV"}:
        return category_text

    return "PV"


def normalize_upper_category(value: Any) -> str:
    text = normalize_text(value).upper()
    if "HP" in text:
        return "HP"
    if "ESS" in text:
        return "ESS"
    if "PV" in text:
        return "PV"
    return text


def map_transit_wh_code(in_transit_code: Any) -> str | None:
    text = normalize_text(in_transit_code).upper()
    if not text:
        return None

    for code, wh in WH_CODE_MAP.items():
        if code in text:
            return wh
    return None


def excluded_row(brand: Any, sales_org_name: Any, virtual_warehouse_name: Any) -> bool:
    if normalize_lower(brand) == "other":
        return True

    sales_org = normalize_lower(sales_org_name)
    if sales_org in {"china", "business planning department"}:
        return True

    if "arrival plan" in normalize_lower(virtual_warehouse_name):
        return True

    return False


def pick_sku_sheet_name(workbook_sheet_names: list[str], explicit_name: str | None) -> str:
    if explicit_name:
        if explicit_name in workbook_sheet_names:
            return explicit_name
        raise ValueError(f"SKU sheet not found: {explicit_name}")

    lowered = {name.lower(): name for name in workbook_sheet_names}
    for candidate in SKU_SHEET_CANDIDATES:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return workbook_sheet_names[0]


def iter_dates(start_date: date, end_date: date) -> list[date]:
    days: list[date] = []
    d = start_date
    while d <= end_date:
        days.append(d)
        d += timedelta(days=1)
    return days


def date_header(d: date) -> str:
    return f"{d.year}.{d.month}.{d.day}"


def parse_supply_date(value: Any) -> date | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalize_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    text = text.replace(".", "-").replace("/", "-")
    parts = text.split("-")
    if len(parts) >= 3 and all(p.isdigit() for p in parts[:3]):
        return date(int(parts[0]), int(parts[1]), int(parts[2]))

    return None


def merge_transit_payload(
    base_qty: dict[tuple[str, str, str], float],
    base_category: dict[tuple[str, str], str],
    add_qty: dict[tuple[str, str, str], float],
    add_category: dict[tuple[str, str], str],
) -> None:
    for key, qty in add_qty.items():
        base_qty[key] += qty
    for key, category in add_category.items():
        if key not in base_category and category:
            base_category[key] = category


def extract_inventory_rows(inventory_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(inventory_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    idx = build_header_index(headers)

    required = [
        "customer model",
        "category",
        "available stock",
        "sales organization name",
        "virtual warehouse name",
        "brand",
    ]
    missing = [h for h in required if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Inventory Detail missing columns: {missing}")

    grouped: dict[tuple[str, str, str], float] = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = row[idx["brand"]]
        sales_org_name = row[idx["sales organization name"]]
        virtual_warehouse_name = row[idx["virtual warehouse name"]]
        if excluded_row(brand, sales_org_name, virtual_warehouse_name):
            continue

        sku = normalize_text(row[idx["customer model"]])
        category = normalize_text(row[idx["category"]])
        sales_org = normalize_text(sales_org_name)
        if not sku:
            continue

        stock = safe_float(row[idx["available stock"]])
        grouped[(sku, category, sales_org)] += stock

    wb.close()

    out: list[dict[str, Any]] = []
    for (sku, category, sales_org), stock in sorted(grouped.items()):
        out.append(
            {
                "SKU": sku,
                "Category": category,
                "Stock": round(stock, 3),
                "WH": map_country_code(sales_org),
            }
        )
    return out


def build_sku_lookup(stock_wb_path: Path, sku_sheet_name: str) -> dict[str, dict[str, Any]]:
    wb = load_workbook(stock_wb_path, data_only=True)
    ws = wb[sku_sheet_name]
    headers = [cell.value for cell in ws[1]]
    idx = build_header_index(headers)

    required = ["sku no.", "category", "pv category", "bins", "container loading", "product tcl report", "anaplan pl6"]
    missing = [h for h in required if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"SKU sheet missing columns: {missing}")

    lookup: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = normalize_text(row[idx["sku no."]])
        if not sku or sku in lookup:
            continue
        lookup[sku] = {
            "Family": row[idx["category"]],
            "Series": row[idx["pv category"]],
            "Model": row[idx["anaplan pl6"]],
            "Bin": row[idx["bins"]],
            "MOQ": row[idx["container loading"]],
            "Product TCL Report": row[idx["product tcl report"]],
        }

    wb.close()
    return lookup


def extract_transit_data(
    daily_supply_plan_path: Path,
    start_date: date,
    end_date: date,
) -> tuple[dict[tuple[str, str, str], float], dict[tuple[str, str], str]]:
    wb = load_workbook(daily_supply_plan_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    idx = build_header_index(headers)
    required = ["in-transit warehouse(code)", "customer model", "supply date", "available quantity"]
    missing = [h for h in required if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"DailySupplyPlan missing columns: {missing}")

    transit_qty: dict[tuple[str, str, str], float] = defaultdict(float)
    transit_category: dict[tuple[str, str], str] = {}

    has_category = "category" in idx

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = normalize_text(row[idx["customer model"]])
        if not sku:
            continue

        wh = map_transit_wh_code(row[idx["in-transit warehouse(code)"]])
        if not wh:
            continue

        supply_date = parse_supply_date(row[idx["supply date"]])
        if supply_date is None or supply_date < start_date or supply_date > end_date:
            continue

        qty = safe_float(row[idx["available quantity"]])
        d_header = date_header(supply_date)
        transit_qty[(sku, wh, d_header)] += qty

        if has_category:
            category_val = normalize_text(row[idx["category"]])
            if category_val and (sku, wh) not in transit_category:
                transit_category[(sku, wh)] = category_val

    wb.close()
    return transit_qty, transit_category


def extract_odp_transit_data(
    odp_master_path: Path,
    start_date: date,
    end_date: date,
) -> tuple[dict[tuple[str, str, str], float], dict[tuple[str, str], str]]:
    wb = load_workbook(odp_master_path, data_only=True, read_only=True)

    sheet_name = None
    for candidate in ("Total Stcok", "Total Stock"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        wb.close()
        raise ValueError("ODP file missing `Total Stcok`/`Total Stock` sheet")

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    idx = build_header_index(headers)

    required = ["new ark wh", "new ark sku", "quantity", "eta for new ark update"]
    missing = [h for h in required if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"ODP Total Stock missing columns: {missing}")

    transit_qty: dict[tuple[str, str, str], float] = defaultdict(float)
    transit_category: dict[tuple[str, str], str] = {}

    has_product_type = "product type" in idx

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = normalize_text(row[idx["new ark sku"]])
        if not sku:
            continue

        raw_wh = row[idx["new ark wh"]]
        if normalize_lower(raw_wh) in {"", "n/a", "na", "none", "null"}:
            continue
        wh = map_transit_wh_code(raw_wh)
        if not wh:
            continue

        raw_eta = row[idx["eta for new ark update"]]
        if normalize_lower(raw_eta) in {"", "n/a", "na", "none", "null"}:
            continue
        eta_date = parse_supply_date(raw_eta)
        if eta_date is None or eta_date.year <= 1900:
            continue
        if eta_date < start_date or eta_date > end_date:
            continue

        qty = safe_float(row[idx["quantity"]])
        d_header = date_header(eta_date)
        transit_qty[(sku, wh, d_header)] += qty

        if has_product_type:
            category_val = normalize_text(row[idx["product type"]])
            if category_val and (sku, wh) not in transit_category:
                transit_category[(sku, wh)] = category_val

    wb.close()
    return transit_qty, transit_category


def extract_to_be_allocated_orders(
    order_file_path: Path,
) -> tuple[list[dict[str, Any]], dict[tuple[str, str], float]]:
    wb = load_workbook(order_file_path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]

    header_row = 2
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    idx = build_header_index(headers)

    required = [
        "allocation status",
        "material",
        "ordered qty",
        "crd",
        "customer level 6 name",
        "so no.",
        "so line",
        "model",
        "factory",
    ]
    missing = [h for h in required if h not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Order file missing columns: {missing}")

    orders: list[dict[str, Any]] = []
    agg_by_sku_wh: dict[tuple[str, str], float] = defaultdict(float)

    max_row = ws.max_row
    for r in range(header_row + 1, max_row + 1):
        status = normalize_lower(ws.cell(r, idx["allocation status"] + 1).value)
        if status != "to be allocated":
            continue

        sku = normalize_text(ws.cell(r, idx["material"] + 1).value)
        qty = safe_float(ws.cell(r, idx["ordered qty"] + 1).value)
        crd = ws.cell(r, idx["crd"] + 1).value
        customer_name = ws.cell(r, idx["customer level 6 name"] + 1).value
        so_no = ws.cell(r, idx["so no."] + 1).value
        so_line = ws.cell(r, idx["so line"] + 1).value
        model = ws.cell(r, idx["model"] + 1).value
        factory = ws.cell(r, idx["factory"] + 1).value
        wh = map_transit_wh_code(factory) or ""

        orders.append(
            {
                "SKU": sku,
                "Ordered Qty": round(qty, 3),
                "CRD": crd,
                "Customer Name": customer_name,
                "SO No.": so_no,
                "SO Line": so_line,
                "Model": model,
                "Factory": factory,
                "WH": wh,
            }
        )

        if sku and wh:
            agg_by_sku_wh[(sku, wh)] += qty

    wb.close()
    return orders, agg_by_sku_wh


def write_output_sheet(
    stock_wb_path: Path,
    rows: list[dict[str, Any]],
    sku_lookup: dict[str, dict[str, Any]],
    transit_qty: dict[tuple[str, str, str], float] | None = None,
    transit_category: dict[tuple[str, str], str] | None = None,
    allocated_orders: list[dict[str, Any]] | None = None,
    allocated_need: dict[tuple[str, str], float] | None = None,
    start_date: date = TRANSIT_START_DATE,
    end_date: date = TRANSIT_END_DATE,
) -> dict[str, Any]:
    wb = load_workbook(stock_wb_path)

    if OUTPUT_SHEET_NAME in wb.sheetnames:
        ws_old = wb[OUTPUT_SHEET_NAME]
        wb.remove(ws_old)
    if ALLOC_SHEET_NAME in wb.sheetnames:
        alloc_old = wb[ALLOC_SHEET_NAME]
        wb.remove(alloc_old)
    ws = wb.create_sheet(OUTPUT_SHEET_NAME)

    base_headers = [
        "WH",
        "Category",
        "Product TCL Report",
        "Family",
        "Series",
        "SKU",
        "Model",
        "Bin",
        "MOQ",
        "To be allocated",
        "Total QTY",
        "Total MW",
        "MW",
        "Stock",
    ]
    transit_dates = iter_dates(start_date, end_date)
    transit_headers = [date_header(d) for d in transit_dates]
    headers = base_headers + transit_headers
    ws.append(headers)

    for item in rows:
        mapped = sku_lookup.get(item["SKU"], {})
        ws.append(
            [
                item["WH"],
                resolve_output_category(item.get("Category"), mapped.get("Family")),
                mapped.get("Product TCL Report"),
                mapped.get("Family"),
                mapped.get("Series"),
                item["SKU"],
                mapped.get("Model"),
                mapped.get("Bin"),
                mapped.get("MOQ"),
                0,
                0,
                0,
                0,
                item["Stock"],
            ]
            + [None] * len(transit_headers)
        )

    header_col = {h: i + 1 for i, h in enumerate(headers)}
    row_by_key: dict[tuple[str, str], int] = {}
    for r in range(2, ws.max_row + 1):
        sku = normalize_text(ws.cell(r, header_col["SKU"]).value)
        wh = normalize_text(ws.cell(r, header_col["WH"]).value)
        if sku and wh and (sku, wh) not in row_by_key:
            row_by_key[(sku, wh)] = r

    transit_qty = transit_qty or {}
    transit_category = transit_category or {}
    allocated_orders = allocated_orders or []
    allocated_need = allocated_need or {}

    for (sku, wh, d_header), qty in transit_qty.items():
        if d_header not in header_col:
            continue

        key = (sku, wh)
        if key not in row_by_key:
            mapped = sku_lookup.get(sku, {})
            category = transit_category.get(key, "")
            resolved_category = resolve_output_category(category, mapped.get("Family"))
            ws.append(
                [
                    wh,
                    resolved_category,
                    mapped.get("Product TCL Report"),
                    mapped.get("Family"),
                    mapped.get("Series"),
                    sku,
                    mapped.get("Model"),
                    mapped.get("Bin"),
                    mapped.get("MOQ"),
                    0,
                    0,
                    0,
                    0,
                    0,
                ]
                + [None] * len(transit_headers)
            )
            row_by_key[key] = ws.max_row

        row_num = row_by_key[key]
        col_num = header_col[d_header]
        current = safe_float(ws.cell(row_num, col_num).value)
        ws.cell(row_num, col_num).value = round(current + qty, 3)

    for r in range(2, ws.max_row + 1):
        sku_val = normalize_text(ws.cell(r, header_col["SKU"]).value)
        wh_val = normalize_text(ws.cell(r, header_col["WH"]).value)
        need_qty = safe_float(allocated_need.get((sku_val, wh_val), 0))
        ws.cell(r, header_col["To be allocated"]).value = round(need_qty, 3)

        bin_qty = safe_float(ws.cell(r, header_col["Bin"]).value)
        stock_qty = safe_float(ws.cell(r, header_col["Stock"]).value)

        transit_total = 0.0
        for d_col_name in transit_headers:
            transit_total += safe_float(ws.cell(r, header_col[d_col_name]).value)

        total_qty = stock_qty + transit_total
        mw = (stock_qty * bin_qty) / 1_000_000
        total_mw = (total_qty * bin_qty) / 1_000_000

        ws.cell(r, header_col["Total QTY"]).value = round(total_qty, 3)
        ws.cell(r, header_col["MW"]).value = round(mw, 3)
        ws.cell(r, header_col["Total MW"]).value = round(total_mw, 3)

    before_cleanup_rows = ws.max_row - 1
    deleted_rows_info: list[tuple[int, str, str, list[str]]] = []
    for r in range(2, ws.max_row + 1):
        sku_val = normalize_text(ws.cell(r, header_col["SKU"]).value)
        wh_val = normalize_text(ws.cell(r, header_col["WH"]).value)
        model_val = normalize_text(ws.cell(r, header_col["Model"]).value)

        reasons: list[str] = []
        if not model_val:
            reasons.append("Model empty")
        if sku_val and sku_val not in sku_lookup:
            reasons.append("SKU not matched")

        if reasons:
            deleted_rows_info.append((r, sku_val, wh_val, reasons))

    for r, _, _, _ in reversed(deleted_rows_info):
        ws.delete_rows(r, 1)

    deleted_model_empty_count = sum(1 for _, _, _, reasons in deleted_rows_info if "Model empty" in reasons)
    deleted_unmatched_sku_count = sum(1 for _, _, _, reasons in deleted_rows_info if "SKU not matched" in reasons)

    deleted_row_samples: list[dict[str, Any]] = []
    for _, sku_val, wh_val, reasons in deleted_rows_info[:20]:
        deleted_row_samples.append(
            {
                "SKU": sku_val,
                "WH": wh_val,
                "Reason": "; ".join(reasons),
            }
        )

    ws_alloc = wb.create_sheet(ALLOC_SHEET_NAME)
    ws_alloc.append(["SKU", "Ordered Qty", "CRD", "Customer Name", "SO No.", "SO Line", "Model", "Factory", "WH"])
    for order in allocated_orders:
        ws_alloc.append(
            [
                order.get("SKU"),
                order.get("Ordered Qty"),
                order.get("CRD"),
                order.get("Customer Name"),
                order.get("SO No."),
                order.get("SO Line"),
                order.get("Model"),
                order.get("Factory"),
                order.get("WH"),
            ]
        )

    wb.save(stock_wb_path)
    wb.close()

    return {
        "before_cleanup_rows": before_cleanup_rows,
        "after_cleanup_rows": ws.max_row - 1,
        "deleted_rows": len(deleted_rows_info),
        "deleted_model_empty_rows": deleted_model_empty_count,
        "deleted_unmatched_sku_rows": deleted_unmatched_sku_count,
        "deleted_row_samples": deleted_row_samples,
    }


def run(
    inventory_path: Path,
    stock_path: Path,
    sku_sheet_name: str | None = None,
    daily_supply_plan_path: Path | None = None,
    odp_master_path: Path | None = None,
    order_file_path: Path | None = None,
    transit_start_date: date = TRANSIT_START_DATE,
    transit_end_date: date = TRANSIT_END_DATE,
) -> None:
    inventory_rows = extract_inventory_rows(inventory_path)

    wb_probe = load_workbook(stock_path, read_only=True)
    resolved_sku_sheet = pick_sku_sheet_name(wb_probe.sheetnames, sku_sheet_name)
    wb_probe.close()

    sku_lookup = build_sku_lookup(stock_path, resolved_sku_sheet)

    transit_qty: dict[tuple[str, str, str], float] = defaultdict(float)
    transit_category: dict[tuple[str, str], str] = {}

    if daily_supply_plan_path:
        d_qty, d_category = extract_transit_data(
            daily_supply_plan_path,
            start_date=transit_start_date,
            end_date=transit_end_date,
        )
        merge_transit_payload(transit_qty, transit_category, d_qty, d_category)

    if odp_master_path:
        o_qty, o_category = extract_odp_transit_data(
            odp_master_path,
            start_date=transit_start_date,
            end_date=transit_end_date,
        )
        merge_transit_payload(transit_qty, transit_category, o_qty, o_category)

    allocated_orders: list[dict[str, Any]] = []
    allocated_need: dict[tuple[str, str], float] = defaultdict(float)
    if order_file_path:
        allocated_orders, allocated_need = extract_to_be_allocated_orders(order_file_path)

    run_summary = write_output_sheet(
        stock_path,
        inventory_rows,
        sku_lookup,
        transit_qty=transit_qty,
        transit_category=transit_category,
        allocated_orders=allocated_orders,
        allocated_need=allocated_need,
        start_date=transit_start_date,
        end_date=transit_end_date,
    )

    print(f"Done: wrote {stock_path} -> sheet {OUTPUT_SHEET_NAME}")
    print(f"Inventory grouped rows: {len(inventory_rows)}")
    print(f"SKU mapping sheet: {resolved_sku_sheet}")
    if daily_supply_plan_path:
        print(f"Transit source (DailySupplyPlan): {daily_supply_plan_path}")
    if odp_master_path:
        print(f"Transit source (ODP Total Stock): {odp_master_path}")
    if order_file_path:
        print(f"Order source (To be allocated): {order_file_path}")
    if daily_supply_plan_path or odp_master_path:
        print(f"Transit date columns: {date_header(transit_start_date)} ~ {date_header(transit_end_date)}")

    print(
        "Rows removed from stock sheet: "
        f"total={run_summary['deleted_rows']}, "
        f"model_empty={run_summary['deleted_model_empty_rows']}, "
        f"sku_unmatched={run_summary['deleted_unmatched_sku_rows']}"
    )

    log_lines: list[str] = [
        f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Output workbook: {stock_path}",
        f"Output sheet: {OUTPUT_SHEET_NAME}",
        f"Inventory grouped rows: {len(inventory_rows)}",
        f"SKU mapping sheet: {resolved_sku_sheet}",
        f"Rows before cleanup: {run_summary['before_cleanup_rows']}",
        f"Rows after cleanup: {run_summary['after_cleanup_rows']}",
        f"Rows removed (total): {run_summary['deleted_rows']}",
        f"Rows removed (Model empty): {run_summary['deleted_model_empty_rows']}",
        f"Rows removed (SKU unmatched): {run_summary['deleted_unmatched_sku_rows']}",
    ]

    if daily_supply_plan_path:
        log_lines.append(f"Transit source (DailySupplyPlan): {daily_supply_plan_path}")
    if odp_master_path:
        log_lines.append(f"Transit source (ODP Total Stock): {odp_master_path}")
    if order_file_path:
        log_lines.append(f"Order source (To be allocated): {order_file_path}")
    if daily_supply_plan_path or odp_master_path:
        log_lines.append(f"Transit date columns: {date_header(transit_start_date)} ~ {date_header(transit_end_date)}")

    samples = run_summary.get("deleted_row_samples", [])
    if samples:
        log_lines.append("Deleted row samples (max 20):")
        for idx, sample in enumerate(samples, start=1):
            log_lines.append(
                f"  {idx}. SKU={sample.get('SKU', '')}, WH={sample.get('WH', '')}, Reason={sample.get('Reason', '')}"
            )
    else:
        log_lines.append("Deleted row samples: none")

    log_path = stock_path.with_name(
        f"{stock_path.stem}_run_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    )
    log_path.write_text("\n".join(log_lines), encoding="utf-8")
    print(f"Run log written: {log_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build stock sheet from Inventory Details + transit plan")
    parser.add_argument("--inventory", required=True, type=Path, help="Inventory Details path")
    parser.add_argument("--stock", required=True, type=Path, help="Stock workbook path")
    parser.add_argument("--sku-sheet", default=None, help="SKU mapping sheet name (optional)")
    parser.add_argument("--daily-supply-plan", default=None, type=Path, help="DailySupplyPlan path (optional)")
    parser.add_argument("--odp-master", default=None, type=Path, help="EUPV_ODP_MASTER path (optional)")
    parser.add_argument("--order-file", default=None, type=Path, help="Orderfile_Base_Realtime path (optional)")
    parser.add_argument("--transit-start", default="2026-08-01", help="Transit start date YYYY-MM-DD")
    parser.add_argument("--transit-end", default="2026-12-31", help="Transit end date YYYY-MM-DD")
    return parser.parse_args()


def _parse_cli_date(text: str) -> date:
    try:
        return datetime.strptime(text.strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"Invalid date: {text}, expected YYYY-MM-DD") from exc


if __name__ == "__main__":
    args = parse_args()
    run(
        inventory_path=args.inventory,
        stock_path=args.stock,
        sku_sheet_name=args.sku_sheet,
        daily_supply_plan_path=args.daily_supply_plan,
        odp_master_path=args.odp_master,
        order_file_path=args.order_file,
        transit_start_date=_parse_cli_date(args.transit_start),
        transit_end_date=_parse_cli_date(args.transit_end),
    )
